"""
Mustering Service - Core Business Logic for POB v2.0
Complete mustering system with real-time headcount tracking
"""

from sqlalchemy.orm import Session
from sqlalchemy import and_, func, or_, text, update
from datetime import datetime, timedelta
from typing import List, Dict, Any, Optional
import logging
import json

from app.models.biotime_models import (
    MusteringEvent, MusteringLog, MusteringExpected,
    MusteringSearchSweep, MusteringEscalationRecord,
    PersonnelEmployee, IClockTerminal, AccDoor, AuthUser, PersonnelDepartment,
    IClockTransaction,
)
from app.models.zone import Zone
from app.models.personnel import Personnel

logger = logging.getLogger(__name__)

# PostgreSQL advisory-lock IDs — keep globally unique across the codebase.
# 42 = global guard ensuring only one mustering event is started at a time.
_MUSTER_START_LOCK_ID = 42

class MusteringService:
    """Core mustering business logic service"""
    
    def __init__(self, db: Session):
        self.db = db

    def _adjust_event_counters(
        self,
        event_id: int,
        delta_safe: int = 0,
        delta_missing: int = 0,
        delta_injured: int = 0,
    ) -> None:
        """Atomically adjust mustering event counters in a single SQL UPDATE.

        Using server-side arithmetic prevents the read-modify-write race condition
        that occurs when multiple biometric punches arrive concurrently and each
        reader increments the Python object before the previous commit lands.
        """
        self.db.execute(
            update(MusteringEvent)
            .where(MusteringEvent.id == event_id)
            .values(
                total_safe=MusteringEvent.total_safe + delta_safe,
                total_missing=MusteringEvent.total_missing + delta_missing,
                total_injured=MusteringEvent.total_injured + delta_injured,
            )
        )

    def start_mustering_event(
        self,
        zone_ids: List[int],
        event_type: int,
        initiated_by: int,
        muster_zone_id: int = None,
        notify_sms: bool = False,
        notify_email: bool = False,
        notify_whatsapp: bool = False,
        notify_siren: bool = False,
        notes: str = None
    ) -> Dict[str, Any]:
        """
        Start a mustering event covering one or more access-control zones.

        zone_ids       = AFFECTED/source zones — expected personnel are everyone whose
                         current_zone_id is in them (tracked by access control live).
        muster_zone_id = the target assembly point (a MUSTER_POINT zone) people report
                         TO; its Horus H1 ADMS reader is what marks them safe.

        Event Types: 0=Real, 1=Drill, 2=Fire, 3=Gas, 4=ManDown
        """
        try:
            if not zone_ids:
                raise ValueError("At least one zone must be selected")

            # A muster point is a safe destination, never a source zone — drop any that
            # slipped into the affected list so they don't inflate the expected roster.
            muster_ids = {
                r[0] for r in self.db.query(Zone.id).filter(Zone.zone_type == "MUSTER_POINT").all()
            }
            zone_ids = [z for z in zone_ids if z not in muster_ids]
            if not zone_ids:
                raise ValueError("Select at least one operational (non-muster) affected zone")

            # Validate zones exist
            zones = self.db.query(Zone).filter(Zone.id.in_(zone_ids)).all()
            found_ids = {z.id for z in zones}
            missing = set(zone_ids) - found_ids
            if missing:
                raise ValueError(f"Zone(s) not found: {missing}")

            # Validate the target muster point (if given) really is a muster point.
            if muster_zone_id is not None:
                if muster_zone_id not in muster_ids:
                    raise ValueError("muster_zone_id must be a MUSTER_POINT zone")

            # Acquire the global mustering advisory lock before checking for an active
            # event. The system intentionally allows only ONE active mustering event at a
            # time (see the active_event check below), so a single global lock is correct:
            # it prevents two concurrent requests both finding no active event and both
            # inserting (the check-then-insert race). pg_try_advisory_xact_lock returns
            # FALSE if another session holds the lock; it auto-releases at transaction end.
            #
            # ADVISORY LOCK REGISTRY (keep unique across the codebase):
            #   _MUSTER_START_LOCK_ID = 42  → global single-active-mustering-event guard
            lock_acquired = self.db.execute(
                text("SELECT pg_try_advisory_xact_lock(:lock_id)"),
                {"lock_id": _MUSTER_START_LOCK_ID},
            ).scalar()
            if not lock_acquired:
                raise ValueError("Another mustering event is being started. Please try again in a moment.")

            active_event = self.db.query(MusteringEvent).filter(MusteringEvent.status == 0).first()
            if active_event:
                raise ValueError("An active mustering event already exists. End it before starting a new one.")

            # Calculate expected personnel from zone occupancy (access-control)
            expected_personnel = self._calculate_expected_personnel(zone_ids)
            total_expected = len(expected_personnel)

            # Use first zone as the FK anchor; all zones stored in zone_ids JSONB
            primary_zone_id = zone_ids[0]

            event = MusteringEvent(
                zone_id=primary_zone_id,
                zone_ids=zone_ids,
                muster_zone_id=muster_zone_id,
                event_type=event_type,
                start_time=datetime.utcnow(),
                status=0,
                initiated_by=initiated_by,
                total_expected=total_expected,
                total_safe=0,
                total_missing=total_expected,
                total_injured=0,
                notes=notes,
            )

            self.db.add(event)
            self.db.flush()

            for emp in expected_personnel:
                dept_id = emp.get('dept_id')
                if dept_id is not None:
                    from app.models.biotime_models import PersonnelDepartment
                    dept_exists = self.db.query(PersonnelDepartment.id).filter(
                        PersonnelDepartment.id == dept_id
                    ).first()
                    if not dept_exists:
                        dept_id = None

                self.db.add(MusteringExpected(
                    event_id=event.id,
                    emp_code=emp['emp_code'],
                    emp_name=emp['emp_name'],
                    dept_id=dept_id,
                    shift_id=None,
                    last_punch_time=None,
                    last_punch_area=emp.get('current_zone_name'),
                ))

                self.db.add(MusteringLog(
                    event_id=event.id,
                    emp_code=emp['emp_code'],
                    emp_name=emp['emp_name'],
                    dept_name=emp.get('dept_name'),
                    check_time=datetime.utcnow(),
                    status=0,  # Missing
                    last_punch_area=emp.get('current_zone_name'),
                ))

            # Commit the event and expected roster first so that reader activation
            # failures cannot roll back the event. Personnel accountability matters
            # more than device commands — a muster must start even if a reader's
            # serial number is misconfigured.
            self.db.commit()

            # Reset zone TILE counts for all non-muster-point zones so zone tiles
            # correctly show 0 (everyone should be evacuating to muster points).
            #
            # IMPORTANT: We deliberately do NOT clear Personnel.current_zone_id.
            # That FK is the rescue intelligence — if someone does not scan at a
            # muster point, their current_zone_id still points to the last zone
            # they badged into, which is where a rescue team should search first.
            # The last_punch_area field in MusteringLog already snapshots the zone
            # name at event start for historical accuracy; current_zone_id gives
            # the live queryable FK for grouping missing persons by search zone.
            try:
                muster_zone_ids = [
                    row[0] for row in self.db.query(Zone.id).filter(
                        Zone.zone_type == 'MUSTER_POINT'
                    ).all()
                ]
                zone_filter = Zone.current_occupancy > 0
                if muster_zone_ids:
                    zone_filter = and_(zone_filter, ~Zone.id.in_(muster_zone_ids))
                self.db.query(Zone).filter(zone_filter).update(
                    {'current_occupancy': 0, 'current_personnel_count': 0},
                    synchronize_session=False
                )
                self.db.commit()
                logger.info(f"Zone tile counts reset for muster event {event.id} "
                            f"(Personnel.current_zone_id preserved for rescue tracking)")
            except Exception as occ_err:
                logger.warning(f"Could not reset zone tile counts: {occ_err}")
                try:
                    self.db.rollback()
                except Exception:
                    pass

            # Activate the muster-point readers (best-effort). Directed muster → only
            # the chosen assembly point's reader; otherwise every muster point so any
            # of them can accept check-ins. Muster readers live on the MUSTER_POINT
            # zone, not the affected/source zones.
            reader_zone_ids = [muster_zone_id] if muster_zone_id else list(muster_ids)
            for zid in reader_zone_ids:
                try:
                    self._set_mustering_readers(zid, True)
                except Exception as reader_err:
                    logger.warning(
                        f"Could not activate mustering reader for muster zone {zid}: {reader_err} — "
                        f"event already committed, continuing"
                    )
                    # Reset the session in case the transaction is now aborted
                    try:
                        self.db.rollback()
                    except Exception:
                        pass

            zone_names = [z.name for z in zones]
            logger.info(
                "Mustering event %d started — zones: %s — %d expected",
                event.id, zone_names, total_expected
            )

            return {
                "event_id": event.id,
                "zone_ids": zone_ids,
                "zone_names": zone_names,
                "event_type": event_type,
                "total_expected": total_expected,
                "start_time": event.start_time.isoformat(),
                "status": "started",
            }

        except Exception as e:
            logger.error(f"Error starting mustering event: {e}")
            self.db.rollback()
            raise
    
    def end_mustering_event(self, event_id: int, ended_by: int, reason: str = None) -> Dict[str, Any]:
        """End a mustering event"""
        try:
            event = self.db.query(MusteringEvent).filter(MusteringEvent.id == event_id).first()
            if not event:
                raise ValueError(f"Event {event_id} not found")
            
            if event.status != 0:
                raise ValueError(f"Event {event_id} is not active")

            # Close the event FIRST and commit. process_mustering_punch() only
            # attributes punches to events with status==0, so once this commits no
            # new punch can change the counts. Computing the final headcount BEFORE
            # the flip (the previous order) left a race window where a punch landing
            # between the read and the commit was silently overwritten by the
            # absolute-value write below.
            event.status = 1  # Completed
            event.end_time = datetime.utcnow()
            self.db.commit()

            # Now that the event is closed, compute the authoritative final headcount
            # and persist it. No concurrent punch can race this read anymore.
            headcount = self.get_event_headcount(event_id)
            event.total_safe = headcount['total_safe']
            event.total_missing = headcount['total_missing']
            event.total_injured = headcount['total_injured']
            self.db.commit()

            # Clear zone occupancy counters — after a muster event the accumulated
            # counts during the emergency are no longer meaningful; readers will
            # repopulate them as personnel badge back through zone gates.
            try:
                self.db.query(Zone).filter(Zone.current_occupancy > 0).update(
                    {'current_occupancy': 0, 'current_personnel_count': 0},
                    synchronize_session=False,
                )
                self.db.commit()
                logger.info(f"Zone occupancy counters cleared after event {event_id} ended")
            except Exception as clr_err:
                logger.warning(f"Could not clear zone occupancy after event end: {clr_err}")
                try:
                    self.db.rollback()
                except Exception:
                    pass

            # Reset mustering readers (best-effort — bad serial numbers must not
            # prevent the event from being marked complete). Deactivate the same
            # muster-point readers that were activated at start: the chosen target,
            # or every muster point when the event was open (no target).
            if event.muster_zone_id:
                reset_zone_ids = [event.muster_zone_id]
            else:
                reset_zone_ids = [
                    r[0] for r in self.db.query(Zone.id).filter(Zone.zone_type == "MUSTER_POINT").all()
                ]
            for zid in reset_zone_ids:
                try:
                    self._set_mustering_readers(zid, False)
                except Exception as reader_err:
                    logger.warning(
                        f"Could not deactivate mustering reader for muster zone {zid}: {reader_err} — "
                        f"event already committed as completed"
                    )
                    try:
                        self.db.rollback()
                    except Exception:
                        pass

            logger.info(f"Mustering event {event_id} ended")
            
            return {
                "event_id": event_id,
                "status": "completed",
                "duration": (event.end_time - event.start_time).total_seconds(),
                "final_headcount": headcount,
                "end_time": event.end_time.isoformat(),
            }
            
        except Exception as e:
            logger.error(f"Error ending mustering event: {e}")
            self.db.rollback()
            raise
    
    def process_mustering_punch(
        self, 
        emp_code: str, 
        device_sn: str, 
        check_time: datetime = None
    ) -> Dict[str, Any]:
        """
        Process a punch from a mustering reader
        This is called by ADMS when acc_door.mustering_mode=true
        """
        try:
            if check_time is None:
                check_time = datetime.utcnow()
            
            # Find active mustering event for this device's zone
            device = self.db.query(IClockTerminal).filter(IClockTerminal.sn == device_sn).first()
            if not device:
                raise ValueError(f"Device {device_sn} not found")
            
            # Find zone for this device — reader_sn, then the terminal's assigned
            # zone_id (per-location headcount), then any active event zone.
            zone = self.db.query(Zone).filter(
                Zone.reader_sn == device_sn
            ).first()

            # Prefer the muster reader's OWN assigned zone so multiple Horus handhelds
            # at different muster points each count to their own location.
            if not zone and getattr(device, 'zone_id', None):
                zone = self.db.query(Zone).filter(Zone.id == device.zone_id).first()

            if not zone:
                # Fallback: find any active mustering event and use its zone.
                # This covers mustering-type terminals (device_type=2) where the zone's
                # reader_sn wasn't explicitly configured.
                active_event_any = self.db.query(MusteringEvent).filter(
                    MusteringEvent.status == 0
                ).first()
                if active_event_any:
                    zone = active_event_any.zone

            if not zone:
                raise ValueError(f"No mustering zone found for device {device_sn}")
            
            # Find active event — check both the primary zone_id and the JSONB zone_ids
            # array so that punches at secondary muster stations are correctly captured.
            event = self.db.query(MusteringEvent).filter(
                and_(
                    or_(
                        MusteringEvent.zone_id == zone.id,
                        MusteringEvent.zone_ids.contains([zone.id]),
                    ),
                    MusteringEvent.status == 0  # Active
                )
            ).first()

            if not event:
                raise ValueError(f"No active mustering event for zone {zone.id}")
            
            # Update or create mustering log
            existing_log = self.db.query(MusteringLog).filter(
                and_(
                    MusteringLog.event_id == event.id,
                    MusteringLog.emp_code == emp_code
                )
            ).first()
            
            if existing_log:
                # Update existing log
                old_status = existing_log.status
                existing_log.check_time = check_time
                existing_log.device_sn = device_sn
                existing_log.device_alias = device.alias
                existing_log.status = 1  # Safe
                
                # Update event counts atomically
                if old_status == 0:  # Was missing
                    self._adjust_event_counters(event.id, delta_safe=+1, delta_missing=-1)
                elif old_status == 2:  # Was injured
                    self._adjust_event_counters(event.id, delta_safe=+1, delta_injured=-1)
                    
            else:
                # Create new log
                employee = self.db.query(PersonnelEmployee).filter(
                    PersonnelEmployee.emp_code == emp_code
                ).first()
                
                log = MusteringLog(
                    event_id=event.id,
                    emp_code=emp_code,
                    emp_name=f"{employee.first_name or ''} {employee.last_name}".strip() if employee else emp_code,
                    dept_name=employee.department.dept_name if employee and employee.department else None,
                    check_time=check_time,
                    device_sn=device_sn,
                    device_alias=device.alias,
                    status=1  # Safe
                )
                self.db.add(log)
                
                # Update event counts atomically
                self._adjust_event_counters(event.id, delta_safe=+1, delta_missing=-1)

            self.db.commit()

            logger.info(f"Processed mustering punch for {emp_code} at {device_sn}")
            
            return {
                "event_id": event.id,
                "emp_code": emp_code,
                "status": "safe",
                "check_time": check_time
            }
            
        except Exception as e:
            logger.error(f"Error processing mustering punch: {e}")
            self.db.rollback()
            raise
    
    def get_event_headcount(self, event_id: int) -> Dict[str, Any]:
        """Get real-time headcount for an event"""
        try:
            event = self.db.query(MusteringEvent).filter(MusteringEvent.id == event_id).first()
            if not event:
                raise ValueError(f"Event {event_id} not found")

            logs = self.db.query(MusteringLog).filter(MusteringLog.event_id == event_id).all()

            total_safe = len([log for log in logs if log.status == 1])
            total_missing = len([log for log in logs if log.status == 0])
            total_injured = len([log for log in logs if log.status == 2])
            total_accounted = total_safe + total_injured

            completion_percentage = (total_accounted / event.total_expected * 100) if event.total_expected > 0 else 0

            # Per-zone breakdown: group logs by the zone each person was in at event start
            zone_tally: dict = {}
            for log in logs:
                area = log.last_punch_area or 'Unknown'
                if area not in zone_tally:
                    zone_tally[area] = {'zone_name': area, 'zone_id': None, 'total': 0, 'safe': 0, 'missing': 0, 'injured': 0}
                zone_tally[area]['total'] += 1
                if log.status == 1:
                    zone_tally[area]['safe'] += 1
                elif log.status == 0:
                    zone_tally[area]['missing'] += 1
                else:
                    zone_tally[area]['injured'] += 1

            # Attach zone IDs for front-end linking
            event_zone_ids = event.zone_ids or ([event.zone_id] if event.zone_id else [])
            if event_zone_ids:
                zone_rows = self.db.query(Zone).filter(Zone.id.in_(event_zone_ids)).all()
                name_to_id = {z.name: z.id for z in zone_rows}
                for entry in zone_tally.values():
                    entry['zone_id'] = name_to_id.get(entry['zone_name'])

            zone_breakdown = sorted(zone_tally.values(), key=lambda x: x['missing'], reverse=True)

            return {
                "event_id": event_id,
                "zone_id": event.zone_id,
                "zone_ids": event_zone_ids,
                "total_expected": event.total_expected,
                "total_safe": total_safe,
                "total_missing": total_missing,
                "total_injured": total_injured,
                "total_accounted": total_accounted,
                "completion_percentage": round(completion_percentage, 2),
                "zone_breakdown": zone_breakdown,
                "last_updated": datetime.utcnow().isoformat()
            }

        except Exception as e:
            logger.error(f"Error getting event headcount: {e}")
            raise
    
    def get_active_events(self) -> List[Dict[str, Any]]:
        """Get all active mustering events"""
        try:
            events = self.db.query(MusteringEvent).filter(MusteringEvent.status == 0).all()
            
            result = []
            for event in events:
                headcount = self.get_event_headcount(event.id)
                result.append({
                    "id": event.id,
                    "zone_id": event.zone_id,
                    "zone_name": event.zone.name if event.zone else None,
                    "event_type": event.event_type,
                    "start_time": event.start_time.isoformat() if event.start_time else None,
                    "initiated_by": event.initiated_by,
                    "headcount": headcount
                })
            
            return result
            
        except Exception as e:
            logger.error(f"Error getting active events: {e}")
            raise
    
    def get_event_logs(self, event_id: int, status: Optional[int] = None) -> List[Dict[str, Any]]:
        """Get logs for a mustering event.

        For MISSING personnel (status=0), the response includes last_known_zone
        (live, from Personnel.current_zone_id) so rescue teams know exactly which
        zone to search. This relies on current_zone_id NOT being cleared at muster
        start — see start_mustering_event for the reasoning.
        """
        try:
            query = (
                self.db.query(MusteringLog, Personnel, Zone)
                .outerjoin(Personnel, Personnel.emp_code == MusteringLog.emp_code)
                .outerjoin(Zone, Zone.id == Personnel.current_zone_id)
                .filter(MusteringLog.event_id == event_id)
            )

            if status is not None:
                query = query.filter(MusteringLog.status == status)

            rows = query.order_by(MusteringLog.check_time.desc()).all()

            result = []
            for log, person, zone in rows:
                is_missing = log.status == 0
                result.append({
                    "id": log.id,
                    "emp_code": log.emp_code,
                    "emp_name": log.emp_name,
                    "dept_name": log.dept_name,
                    "check_time": log.check_time,
                    "device_sn": log.device_sn,
                    "device_alias": getattr(log, 'device_alias', None),
                    "last_punch_area": getattr(log, 'last_punch_area', None),
                    # Rescue intelligence — live zone FK, preserved through the event.
                    # Only meaningful for MISSING personnel (safe/injured are at muster point).
                    "last_known_zone_id": person.current_zone_id if person and is_missing else None,
                    "last_known_zone": zone.name if zone and is_missing else None,
                    "last_known_zone_code": zone.code if zone and is_missing else None,
                    "status": log.status,
                })
            return result

        except Exception as e:
            logger.error(f"Error getting event logs: {e}")
            raise
    
    def export_event_report(self, event_id: int, fmt: str = "excel") -> tuple:
        """
        Build a downloadable report for a completed mustering event.

        Returns (bytes, content_type, filename).
        Supports fmt='excel' (xlsx) and fmt='csv'.
        """
        import io, csv as _csv
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        event = self.db.query(MusteringEvent).filter(MusteringEvent.id == event_id).first()
        if not event:
            raise ValueError(f"Event {event_id} not found")

        logs = (
            self.db.query(MusteringLog)
            .filter(MusteringLog.event_id == event_id)
            .order_by(MusteringLog.status.asc(), MusteringLog.emp_name.asc())
            .all()
        )

        # Resolve zone names
        zone_ids_list = event.zone_ids or ([event.zone_id] if event.zone_id else [])
        zones = self.db.query(Zone).filter(Zone.id.in_(zone_ids_list)).all() if zone_ids_list else []
        zone_label = ", ".join(z.name for z in zones) or "—"

        event_type_map = {0: "Real Emergency", 1: "Drill", 2: "Fire", 3: "Gas Leak", 4: "Man Down"}
        status_map     = {0: "MISSING", 1: "SAFE", 2: "INJURED"}
        ev_type_label  = event_type_map.get(event.event_type, str(event.event_type))

        duration_str = "—"
        if event.end_time and event.start_time:
            secs = int((event.end_time - event.start_time).total_seconds())
            duration_str = f"{secs // 3600:02d}h {(secs % 3600) // 60:02d}m {secs % 60:02d}s"

        total_safe    = sum(1 for l in logs if l.status == 1)
        total_missing = sum(1 for l in logs if l.status == 0)
        total_injured = sum(1 for l in logs if l.status == 2)
        total_exp     = event.total_expected or len(logs)
        pct           = round((total_safe + total_injured) / total_exp * 100, 1) if total_exp else 0

        # ── CSV path ───────────────────────────────────────────────
        if fmt == "csv":
            buf = io.StringIO()
            w = _csv.writer(buf)
            w.writerow(["Mustering Event Report"])
            w.writerow(["Event ID", event_id])
            w.writerow(["Type", ev_type_label])
            w.writerow(["Zones", zone_label])
            w.writerow(["Start", event.start_time.strftime("%Y-%m-%d %H:%M:%S") if event.start_time else ""])
            w.writerow(["End", event.end_time.strftime("%Y-%m-%d %H:%M:%S") if event.end_time else "Ongoing"])
            w.writerow(["Duration", duration_str])
            w.writerow(["Accounted %", f"{pct}%"])
            w.writerow([])
            w.writerow(["#", "Emp Code", "Name", "Department", "Zone Last Seen", "Status", "Checked At"])
            for i, log in enumerate(logs, 1):
                w.writerow([
                    i,
                    log.emp_code,
                    log.emp_name or "",
                    log.dept_name or "",
                    getattr(log, "last_punch_area", "") or "",
                    status_map.get(log.status, "?"),
                    log.check_time.strftime("%Y-%m-%d %H:%M:%S") if log.check_time else "",
                ])
            data = buf.getvalue().encode("utf-8-sig")  # BOM for Excel CSV open
            fname = f"muster_event_{event_id}_{event.start_time.strftime('%Y%m%d')}.csv"
            return data, "text/csv", fname

        # ── Excel path ─────────────────────────────────────────────
        wb = Workbook()

        # Colour palette
        RED    = "FFF5222D"
        GREEN  = "FF52C41A"
        ORANGE = "FFFA8C16"
        BLUE   = "FF1890FF"
        DARK   = "FF141414"
        GREY   = "FFF0F0F0"
        WHITE  = "FFFFFFFF"
        HEADER_FILL = PatternFill("solid", fgColor="FF2D3748")
        MISSING_FILL = PatternFill("solid", fgColor="FFFFF1F0")
        SAFE_FILL    = PatternFill("solid", fgColor="FFF6FFED")
        INJURED_FILL = PatternFill("solid", fgColor="FFFFF7E6")

        thin = Side(style="thin", color="FFD9D9D9")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        def _h(ws, r, c, val, bold=False, size=11, color=DARK, bg=None, align="left"):
            cell = ws.cell(row=r, column=c, value=val)
            cell.font = Font(bold=bold, size=size, color=color)
            cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
            if bg:
                cell.fill = PatternFill("solid", fgColor=bg)
            cell.border = border
            return cell

        # ── Sheet 1: Summary ───────────────────────────────────────
        ws1 = wb.active
        ws1.title = "Summary"
        ws1.column_dimensions["A"].width = 22
        ws1.column_dimensions["B"].width = 38

        title_cell = ws1.cell(row=1, column=1, value="MUSTERING EVENT REPORT")
        title_cell.font = Font(bold=True, size=16, color=WHITE)
        title_cell.fill = PatternFill("solid", fgColor="FF1A202C")
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws1.merge_cells("A1:B1")
        ws1.row_dimensions[1].height = 36

        rows = [
            ("Event ID",       event_id),
            ("Emergency Type", ev_type_label),
            ("Zones",          zone_label),
            ("Start Time",     event.start_time.strftime("%d %b %Y  %H:%M:%S") if event.start_time else "—"),
            ("End Time",       event.end_time.strftime("%d %b %Y  %H:%M:%S") if event.end_time else "Ongoing"),
            ("Duration",       duration_str),
            ("Status",         {0: "Active", 1: "Completed", 2: "Cancelled"}.get(event.status, "?")),
            ("Notes",          event.notes or ""),
        ]
        for i, (label, val) in enumerate(rows, 3):
            _h(ws1, i, 1, label, bold=True, bg="FFEBF0F5")
            _h(ws1, i, 2, val)
            ws1.row_dimensions[i].height = 20

        # Headcount block
        ws1.row_dimensions[11].height = 8
        hc_headers = [("Total Expected", total_exp, BLUE),
                      ("Confirmed Safe", total_safe, GREEN),
                      ("Missing",        total_missing, RED),
                      ("Injured",        total_injured, ORANGE),
                      ("Accounted %",    f"{pct}%",  BLUE)]
        for j, (label, val, col) in enumerate(hc_headers, 1):
            ws1.cell(row=12, column=j).value = label
            ws1.cell(row=12, column=j).font = Font(bold=True, size=9, color=WHITE)
            ws1.cell(row=12, column=j).fill = PatternFill("solid", fgColor=col)
            ws1.cell(row=12, column=j).alignment = Alignment(horizontal="center")
            ws1.cell(row=13, column=j).value = val
            ws1.cell(row=13, column=j).font = Font(bold=True, size=18, color=col)
            ws1.cell(row=13, column=j).alignment = Alignment(horizontal="center")
            ws1.cell(row=13, column=j).fill = PatternFill("solid", fgColor="FFF7FAFC")
            ws1.row_dimensions[13].height = 32
        for j in range(1, 6):
            ws1.column_dimensions[get_column_letter(j)].width = 18

        # ── Sheet 2: Full Roster ───────────────────────────────────
        ws2 = wb.create_sheet("Personnel Roster")
        col_widths = [5, 14, 28, 22, 22, 14, 20]
        for ci, w in enumerate(col_widths, 1):
            ws2.column_dimensions[get_column_letter(ci)].width = w

        headers = ["#", "Emp Code", "Name", "Department", "Zone Last Seen", "Status", "Checked At"]
        for ci, h in enumerate(headers, 1):
            cell = ws2.cell(row=1, column=ci, value=h)
            cell.font = Font(bold=True, size=10, color=WHITE)
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border
        ws2.row_dimensions[1].height = 22

        for i, log in enumerate(logs, 1):
            fill = {0: MISSING_FILL, 1: SAFE_FILL, 2: INJURED_FILL}.get(log.status, PatternFill())
            row_data = [
                i,
                log.emp_code,
                log.emp_name or "",
                log.dept_name or "",
                getattr(log, "last_punch_area", "") or "",
                status_map.get(log.status, "?"),
                log.check_time.strftime("%d %b %Y  %H:%M") if log.check_time else "",
            ]
            for ci, val in enumerate(row_data, 1):
                cell = ws2.cell(row=i + 1, column=ci, value=val)
                cell.fill = fill
                cell.border = border
                cell.alignment = Alignment(vertical="center")
                if ci == 6:  # Status column — bold + color
                    cell.font = Font(bold=True, color={0: RED, 1: GREEN, 2: ORANGE}.get(log.status, DARK))
                    cell.alignment = Alignment(horizontal="center", vertical="center")
            ws2.row_dimensions[i + 1].height = 18

        ws2.freeze_panes = "A2"
        ws2.auto_filter.ref = f"A1:G{len(logs) + 1}"

        # ── Sheet 3: Missing Persons ───────────────────────────────
        missing = [l for l in logs if l.status == 0]
        ws3 = wb.create_sheet("Missing Persons")
        if missing:
            for ci, (h, w) in enumerate(zip(["#","Emp Code","Name","Department","Zone Last Seen","Last Updated"],
                                             [5, 14, 28, 22, 22, 20]), 1):
                ws3.column_dimensions[get_column_letter(ci)].width = w
                cell = ws3.cell(row=1, column=ci, value=h)
                cell.font = Font(bold=True, size=10, color=WHITE)
                cell.fill = PatternFill("solid", fgColor="FFA8071A")
                cell.alignment = Alignment(horizontal="center")
                cell.border = border
            ws3.row_dimensions[1].height = 22
            for i, log in enumerate(missing, 1):
                for ci, val in enumerate([i, log.emp_code, log.emp_name or "", log.dept_name or "",
                                          getattr(log,"last_punch_area","") or "",
                                          log.check_time.strftime("%d %b  %H:%M") if log.check_time else ""], 1):
                    cell = ws3.cell(row=i+1, column=ci, value=val)
                    cell.fill = MISSING_FILL
                    cell.border = border
                    cell.alignment = Alignment(vertical="center")
                ws3.row_dimensions[i+1].height = 18
        else:
            ws3.cell(row=1, column=1, value="✓  All personnel accounted for — no missing persons").font = Font(bold=True, color="FF52C41A", size=12)

        buf = io.BytesIO()
        wb.save(buf)
        fname = f"muster_event_{event_id}_{event.start_time.strftime('%Y%m%d_%H%M')}.xlsx"
        return buf.getvalue(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", fname

    def mark_person_status(self, event_id: int, emp_code: str, status: int, marked_by: int) -> Dict[str, Any]:
        """Manually mark a person's status"""
        try:
            if status not in [0, 1, 2]:  # Missing, Safe, Injured
                raise ValueError("Invalid status. Must be 0 (Missing), 1 (Safe), or 2 (Injured)")
            
            log = self.db.query(MusteringLog).filter(
                and_(
                    MusteringLog.event_id == event_id,
                    MusteringLog.emp_code == emp_code
                )
            ).first()
            
            if not log:
                raise ValueError(f"No log found for employee {emp_code} in event {event_id}")
            
            old_status = log.status
            log.status = status
            log.check_time = datetime.utcnow()

            # Marking someone Safe/Injured means they've been accounted for at the
            # assembly point — attribute them to the event's target muster point so
            # the headcount shows there (the map credits a zone by last_punch_area
            # name). Only when the event has a directed target; open musters keep
            # whatever area the badge/last-seen recorded.
            if status in (1, 2):
                event = self.db.query(MusteringEvent).filter(MusteringEvent.id == event_id).first()
                if event and event.muster_zone_id:
                    mz = self.db.query(Zone).filter(Zone.id == event.muster_zone_id).first()
                    if mz:
                        log.last_punch_area = mz.name

            # Compute counter deltas and apply atomically
            delta_safe = delta_missing = delta_injured = 0
            if old_status == 0 and status == 1:    # Missing -> Safe
                delta_safe, delta_missing = +1, -1
            elif old_status == 0 and status == 2:  # Missing -> Injured
                delta_injured, delta_missing = +1, -1
            elif old_status == 1 and status == 0:  # Safe -> Missing
                delta_safe, delta_missing = -1, +1
            elif old_status == 1 and status == 2:  # Safe -> Injured
                delta_injured, delta_safe = +1, -1
            elif old_status == 2 and status == 0:  # Injured -> Missing
                delta_injured, delta_missing = -1, +1
            elif old_status == 2 and status == 1:  # Injured -> Safe
                delta_safe, delta_injured = +1, -1

            if any([delta_safe, delta_missing, delta_injured]):
                self._adjust_event_counters(
                    event_id,
                    delta_safe=delta_safe,
                    delta_missing=delta_missing,
                    delta_injured=delta_injured,
                )

            self.db.commit()
            
            logger.info(f"Marked {emp_code} as {status} in event {event_id}")
            
            return {
                "event_id": event_id,
                "emp_code": emp_code,
                "old_status": old_status,
                "new_status": status,
                "marked_by": marked_by
            }
            
        except Exception as e:
            logger.error(f"Error marking person status: {e}")
            self.db.rollback()
            raise
    
    def _calculate_expected_personnel(self, zone_ids: List[int]) -> List[Dict[str, Any]]:
        """Return personnel expected to muster for the given work zones.

        Strategy (two-tier):
        1. Primary: personnel whose current_zone_id is in the selected zones — set
           by access-control readers when someone badges through a zone gate.
        2. Fallback: if no zone-assigned personnel are found (readers not yet
           updating current_zone_id, or zone IDs are muster points), fall back to
           ALL personnel with is_onboard=True.  This ensures the expected list is
           never empty on a live offshore installation where people are present but
           readers haven't been fully configured.
        """
        try:
            def _build_rows(filter_clause):
                return (
                    self.db.query(Personnel, PersonnelEmployee, Zone)
                    .outerjoin(PersonnelEmployee, Personnel.emp_code == PersonnelEmployee.emp_code)
                    .outerjoin(Zone, Zone.id == Personnel.current_zone_id)
                    .filter(filter_clause, Personnel.is_active == True)
                    .all()
                )

            rows = _build_rows(Personnel.current_zone_id.in_(zone_ids))

            used_fallback = False
            if not rows:
                logger.warning(
                    "Zones %s: no personnel have current_zone_id set — "
                    "falling back to all is_onboard=True personnel",
                    zone_ids,
                )
                rows = _build_rows(Personnel.is_onboard == True)
                used_fallback = True

            expected = []
            for person, bio_emp, zone in rows:
                emp_code = person.emp_code or (bio_emp.emp_code if bio_emp else None)
                if not emp_code:
                    continue

                dept_id = bio_emp.dept_id if bio_emp else None
                dept_name = (bio_emp.department.dept_name if bio_emp and bio_emp.department else None)

                expected.append({
                    "emp_code": emp_code,
                    "emp_name": person.full_name or (
                        f"{bio_emp.first_name or ''} {bio_emp.last_name or ''}".strip()
                        if bio_emp else emp_code
                    ),
                    "dept_id": dept_id,
                    "dept_name": dept_name,
                    "current_zone_name": zone.name if zone else None,
                })

            logger.info(
                "Zones %s: %d personnel expected to muster (fallback=%s)",
                zone_ids, len(expected), used_fallback,
            )
            return expected

        except Exception as e:
            logger.error(f"Error calculating expected personnel: {e}")
            raise

    def _set_mustering_readers(self, zone_id: int, mustering_mode: bool):
        """Activate/deactivate mustering mode on all relevant readers for a zone.

        Covers two cases:
        1. The zone has an explicit reader_sn configured.
        2. Any terminal with device_type=2 (dedicated mustering reader) — these
           always participate regardless of zone.reader_sn configuration.
        """
        try:
            zone = self.db.query(Zone).filter(Zone.id == zone_id).first()
            if not zone:
                return

            reader_sns = set()
            if zone.reader_sn:
                reader_sns.add(zone.reader_sn)

            # Also include all dedicated mustering terminals (device_type=2)
            mustering_terminals = self.db.query(IClockTerminal).filter(
                IClockTerminal.device_type == 2
            ).all()
            for t in mustering_terminals:
                reader_sns.add(t.sn)

            for reader_sn in reader_sns:
                door = self.db.query(AccDoor).filter(AccDoor.terminal_sn == reader_sn).first()
                if door:
                    door.mustering_mode = mustering_mode
                    logger.info(f"Set reader {reader_sn} mustering_mode={mustering_mode}")
                else:
                    logger.info(f"No AccDoor row for {reader_sn} — skipping door record update")

                command = 1 if mustering_mode else 0
                # Verify the terminal exists in iclock_terminal before inserting a
                # devcmd row — the FK constraint will reject unknown serial numbers.
                # This can happen when a zone.reader_sn points to a deleted or
                # never-registered terminal; skip silently rather than aborting.
                terminal_exists = self.db.query(IClockTerminal).filter(
                    IClockTerminal.sn == reader_sn
                ).first()
                if not terminal_exists:
                    logger.warning(
                        f"Terminal {reader_sn!r} not found in iclock_terminal — "
                        f"skipping SET MUSTERING MODE devcmd"
                    )
                    continue
                self.db.execute(text("""
                    INSERT INTO iclock_devcmd (sn, cmd_content, status, cmd_commit_time)
                    VALUES (:sn, :cmd, 0, :now)
                """), {
                    'sn': reader_sn,
                    'cmd': f"SET MUSTERING MODE={command}",
                    'now': datetime.utcnow(),
                })
                logger.info(f"Queued SET MUSTERING MODE={command} for reader {reader_sn}")

            self.db.commit()

        except Exception as e:
            logger.error(f"Error setting mustering readers: {e}")
            raise
    
    def _trigger_sirens(self, zone_id: int):
        """Trigger emergency sirens in zone"""
        try:
            # Find emergency devices in zone
            # This would query emergency_device table and send EMERGENCY_ON commands
            logger.info(f"Triggering sirens for zone {zone_id}")
            
        except Exception as e:
            logger.error(f"Error triggering sirens: {e}")
            raise
    
    # ── Accountability closure loop ─────────────────────────────────────────────

    def get_missing_with_escalation(self, event_id: int) -> List[Dict[str, Any]]:
        """
        Return all missing persons for an event, enriched with:
        - last known location (last biometric punch)
        - minutes since event start (escalation clock)
        - escalation level: 0=new <10min, 1=alert 10-20min, 2=search 20-30min, 3=critical >30min
        - search sweep history
        """
        event = self.db.query(MusteringEvent).filter(MusteringEvent.id == event_id).first()
        if not event:
            raise ValueError(f"Event {event_id} not found")

        now = datetime.utcnow()
        start = event.start_time
        if hasattr(start, 'tzinfo') and start.tzinfo is not None:
            start = start.replace(tzinfo=None)
        minutes_elapsed = (now - start).total_seconds() / 60

        missing_logs = self.db.query(MusteringLog).filter(
            and_(MusteringLog.event_id == event_id, MusteringLog.status == 0)
        ).all()

        result = []
        for log in missing_logs:
            last_txn = (
                self.db.query(IClockTransaction)
                .filter(IClockTransaction.emp_code == log.emp_code)
                .order_by(IClockTransaction.punch_time.desc())
                .first()
            )

            if minutes_elapsed >= 30:
                level, label, color = 3, 'CRITICAL', '#ff0000'
            elif minutes_elapsed >= 20:
                level, label, color = 2, 'SEARCH ORDERED', '#ff7a00'
            elif minutes_elapsed >= 10:
                level, label, color = 1, 'ALERT', '#faad14'
            else:
                level, label, color = 0, 'MISSING', '#ff4d4f'

            sweeps = (
                self.db.query(MusteringSearchSweep)
                .filter(
                    and_(
                        MusteringSearchSweep.event_id == event_id,
                        MusteringSearchSweep.emp_code == log.emp_code,
                    )
                )
                .order_by(MusteringSearchSweep.sweep_time.desc())
                .all()
            )

            result.append({
                "emp_code": log.emp_code,
                "emp_name": log.emp_name or log.emp_code,
                "dept_name": log.dept_name,
                "minutes_missing": round(minutes_elapsed, 1),
                "escalation_level": level,
                "escalation_label": label,
                "escalation_color": color,
                "last_known_location": (last_txn.area_alias or None) if last_txn else None,
                "last_known_device": last_txn.terminal_sn if last_txn else None,
                "last_seen_at": last_txn.punch_time.isoformat() if last_txn else None,
                "sweep_count": len(sweeps),
                "last_sweep": {
                    "area_searched": sweeps[0].area_searched,
                    "result": sweeps[0].result,
                    "searcher_name": sweeps[0].searcher_name,
                    "sweep_time": sweeps[0].sweep_time.isoformat(),
                } if sweeps else None,
                "sweeps": [
                    {
                        "id": s.id,
                        "area_searched": s.area_searched,
                        "result": s.result,
                        "searcher_name": s.searcher_name,
                        "notes": s.notes,
                        "sweep_time": s.sweep_time.isoformat(),
                    }
                    for s in sweeps
                ],
            })

        result.sort(key=lambda x: (-x["escalation_level"], -x["minutes_missing"]))
        return result

    def record_search_sweep(
        self,
        event_id: int,
        emp_code: str,
        area_searched: str,
        result_status: str,
        searcher_id: int,
        searcher_name: str = None,
        notes: str = None,
    ) -> Dict[str, Any]:
        """Record a manual search sweep for a missing person."""
        valid = {'NOT_FOUND', 'FOUND_SAFE', 'FOUND_INJURED'}
        if result_status not in valid:
            raise ValueError(f"result must be one of {valid}")

        event = self.db.query(MusteringEvent).filter(
            and_(MusteringEvent.id == event_id, MusteringEvent.status == 0)
        ).first()
        if not event:
            raise ValueError(f"No active event {event_id}")

        sweep = MusteringSearchSweep(
            event_id=event_id,
            emp_code=emp_code,
            area_searched=area_searched,
            result=result_status,
            searcher_id=searcher_id,
            searcher_name=searcher_name,
            notes=notes,
        )
        self.db.add(sweep)
        self.db.flush()
        sweep_id = sweep.id

        if result_status == 'FOUND_SAFE':
            self.mark_person_status(event_id, emp_code, 1, searcher_id)
        elif result_status == 'FOUND_INJURED':
            self.mark_person_status(event_id, emp_code, 2, searcher_id)
        else:
            self.db.commit()

        logger.info(f"Search sweep {sweep_id} recorded for {emp_code} in event {event_id}: {result_status}")
        return {
            "id": sweep_id,
            "event_id": event_id,
            "emp_code": emp_code,
            "area_searched": area_searched,
            "result": result_status,
            "searcher_name": searcher_name,
        }

    def get_event_sweeps(self, event_id: int, emp_code: str = None) -> List[Dict[str, Any]]:
        """List all search sweeps for a mustering event, optionally filtered by employee."""
        query = self.db.query(MusteringSearchSweep).filter(
            MusteringSearchSweep.event_id == event_id
        )
        if emp_code:
            query = query.filter(MusteringSearchSweep.emp_code == emp_code)
        sweeps = query.order_by(MusteringSearchSweep.sweep_time.desc()).all()
        return [
            {
                "id": s.id,
                "event_id": s.event_id,
                "emp_code": s.emp_code,
                "area_searched": s.area_searched,
                "result": s.result,
                "searcher_id": s.searcher_id,
                "searcher_name": s.searcher_name,
                "notes": s.notes,
                "sweep_time": s.sweep_time.isoformat(),
            }
            for s in sweeps
        ]

    def _send_notifications(self, event_id: int, sms: bool, email: bool, whatsapp: bool):
        """Send notifications for mustering event"""
        try:
            event = self.db.query(MusteringEvent).filter(MusteringEvent.id == event_id).first()
            if not event:
                return
            
            # Queue notification tasks (in real implementation with Celery)
            if sms:
                logger.info(f"Queueing SMS notifications for event {event_id}")
            
            if email:
                logger.info(f"Queueing email notifications for event {event_id}")
            
            if whatsapp:
                logger.info(f"Queueing WhatsApp notifications for event {event_id}")
            
        except Exception as e:
            logger.error(f"Error sending notifications: {e}")
            raise
