"""
Personnel Export Service

This service handles personnel data export functionality including
CSV, Excel, PDF, and JSON export formats with customizable templates.
"""

from datetime import datetime, timezone
from typing import List, Dict, Any, Optional, Union
from sqlalchemy.orm import Session
from sqlalchemy import and_, or_, func
import csv
import io
import json
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border
from openpyxl.utils import get_column_letter

from ..models.personnel import Personnel, PersonnelStatus
from ..core.database import get_db


class PersonnelExportService:
    """Service for managing personnel data export"""
    
    def __init__(self):
        # Standard export formats
        self.export_formats = {
            "CSV": "Comma-separated values",
            "EXCEL": "Microsoft Excel spreadsheet",
            "JSON": "JavaScript Object Notation",
            "PDF": "Portable Document Format"
        }
        
        # Standard export templates
        self.export_templates = {
            "BASIC": "Basic personnel information",
            "DETAILED": "Complete personnel profile",
            "CONTACTS": "Emergency contacts only",
            "CERTIFICATIONS": "Certification and training records",
            "MEDICAL": "Medical fitness information",
            "AUDIT": "Audit trail and history",
            "COMPLIANCE": "Compliance and status summary",
            "CUSTOM": "Custom field selection"
        }
        
        # Standard field groups
        self.field_groups = {
            "basic": [
                "id", "badge_id", "full_name", "email", "phone", "company", 
                "role", "department", "status", "is_onboard"
            ],
            "detailed": [
                "id", "badge_id", "full_name", "email", "phone", "company", 
                "role", "department", "status", "is_onboard", "blood_group",
                "nationality", "date_of_birth", "gender", "address", "city", 
                "country", "postal_code", "hire_date", "termination_date",
                "created_at", "updated_at"
            ],
            "contacts": [
                "id", "badge_id", "full_name", "emergency_contact"
            ],
            "certifications": [
                "id", "badge_id", "full_name", "certifications"
            ],
            "medical": [
                "id", "badge_id", "full_name", "medical_fitness_date", 
                "emergency_contact"
            ],
            "audit": [
                "id", "badge_id", "full_name", "audit_trail"
            ],
            "compliance": [
                "id", "badge_id", "full_name", "status", "is_onboard", 
                "medical_fitness_date", "certifications", "emergency_contact"
            ]
        }
    
    async def export_personnel_data(
        self,
        export_format: str = "CSV",
        template: str = "BASIC",
        filters: Optional[Dict[str, Any]] = None,
        fields: Optional[List[str]] = None,
        db: Session = None
    ) -> Dict[str, Any]:
        """
        Export personnel data in specified format
        
        Args:
            export_format: Export format (CSV, EXCEL, JSON, PDF)
            template: Export template (BASIC, DETAILED, CONTACTS, etc.)
            filters: Export filters (optional)
            fields: Custom field selection (optional)
            db: Database session
            
        Returns:
            Export result with file data
        """
        if db is None:
            db = next(get_db())
        
        # Validate export format
        if export_format not in self.export_formats:
            raise ValueError(f"Unsupported export format: {export_format}")
        
        # Get personnel data
        personnel_data = await self._get_filtered_personnel(filters, db)
        
        # Process data based on template
        processed_data = await self._process_export_data(
            personnel_data, template, fields, db
        )
        
        # Generate export file
        if export_format == "CSV":
            file_data = await self._generate_csv_export(processed_data)
            file_name = f"personnel_export_{datetime.now(timezone.utc).strftime('%Y%m%d_%H%M%S')}.csv"
            content_type = "text/csv"
        
        elif export_format == "EXCEL":
            file_data = await self._generate_excel_export(processed_data)
            file_name = f"personnel_export_{datetime.now(timezone.utc).strftime('%Y%m%d_%H%M%S')}.xlsx"
            content_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        
        elif export_format == "JSON":
            file_data = await self._generate_json_export(processed_data)
            file_name = f"personnel_export_{datetime.now(timezone.utc).strftime('%Y%m%d_%H%M%S')}.json"
            content_type = "application/json"
        
        elif export_format == "PDF":
            file_data = await self._generate_pdf_export(processed_data)
            file_name = f"personnel_export_{datetime.now(timezone.utc).strftime('%Y%m%d_%H%M%S')}.pdf"
            content_type = "application/pdf"
        
        else:
            raise ValueError(f"Unsupported export format: {export_format}")
        
        return {
            "success": True,
            "file_name": file_name,
            "file_data": file_data,
            "content_type": content_type,
            "record_count": len(processed_data),
            "export_format": export_format,
            "template": template,
            "export_timestamp": datetime.now(timezone.utc).isoformat(),
            "filters_applied": filters or {},
            "fields_exported": fields or self.field_groups.get(template, [])
        }
    
    async def _get_filtered_personnel(
        self,
        filters: Optional[Dict[str, Any]],
        db: Session
    ) -> List[Personnel]:
        """Get filtered personnel data"""
        query = db.query(Personnel)
        
        if filters:
            # Apply filters
            if filters.get("status"):
                query = query.filter(Personnel.status == filters["status"])
            
            if filters.get("is_onboard") is not None:
                query = query.filter(Personnel.is_onboard == filters["is_onboard"])
            
            if filters.get("company"):
                query = query.filter(Personnel.company.ilike(f"%{filters['company']}%"))
            
            if filters.get("department"):
                query = query.filter(Personnel.department.ilike(f"%{filters['department']}%"))
            
            if filters.get("role"):
                query = query.filter(Personnel.role.ilike(f"%{filters['role']}%"))
            
            if filters.get("blood_group"):
                query = query.filter(Personnel.blood_group == filters["blood_group"])
            
            if filters.get("date_from"):
                query = query.filter(Personnel.created_at >= filters["date_from"])
            
            if filters.get("date_to"):
                query = query.filter(Personnel.created_at <= filters["date_to"])
        
        return query.all()
    
    async def _process_export_data(
        self,
        personnel_data: List[Personnel],
        template: str,
        custom_fields: Optional[List[str]],
        db: Session
    ) -> List[Dict[str, Any]]:
        """Process personnel data for export"""
        processed_data = []
        
        # Determine fields to export
        if custom_fields:
            fields = custom_fields
        elif template in self.field_groups:
            fields = self.field_groups[template]
        else:
            fields = self.field_groups["basic"]
        
        for person in personnel_data:
            record = {}
            
            for field in fields:
                if field == "id":
                    record[field] = person.id
                elif field == "badge_id":
                    record[field] = person.badge_id
                elif field == "full_name":
                    record[field] = person.full_name
                elif field == "email":
                    record[field] = person.email
                elif field == "phone":
                    record[field] = person.phone
                elif field == "company":
                    record[field] = person.company
                elif field == "role":
                    record[field] = person.role
                elif field == "department":
                    record[field] = person.department
                elif field == "status":
                    record[field] = person.status.value if person.status else None
                elif field == "is_onboard":
                    record[field] = person.is_onboard
                elif field == "blood_group":
                    record[field] = person.blood_group
                elif field == "nationality":
                    record[field] = person.nationality
                elif field == "date_of_birth":
                    record[field] = person.date_of_birth.isoformat() if person.date_of_birth else None
                elif field == "gender":
                    record[field] = person.gender
                elif field == "address":
                    record[field] = person.address
                elif field == "city":
                    record[field] = person.city
                elif field == "country":
                    record[field] = person.country
                elif field == "postal_code":
                    record[field] = person.postal_code
                elif field == "hire_date":
                    record[field] = person.hire_date.isoformat() if person.hire_date else None
                elif field == "termination_date":
                    record[field] = person.termination_date.isoformat() if person.termination_date else None
                elif field == "created_at":
                    record[field] = person.created_at.isoformat() if person.created_at else None
                elif field == "updated_at":
                    record[field] = person.updated_at.isoformat() if person.updated_at else None
                elif field == "emergency_contact":
                    record[field] = json.dumps(person.emergency_contact) if person.emergency_contact else None
                elif field == "certifications":
                    record[field] = json.dumps(person.certifications) if person.certifications else None
                elif field == "medical_fitness_date":
                    record[field] = person.medical_fitness_date.isoformat() if person.medical_fitness_date else None
                elif field == "audit_trail":
                    record[field] = json.dumps(getattr(person, 'audit_trail', []))
                elif hasattr(person, field):
                    # Handle dynamic attributes
                    value = getattr(person, field)
                    if hasattr(value, 'isoformat'):
                        record[field] = value.isoformat()
                    elif isinstance(value, (dict, list)):
                        record[field] = json.dumps(value)
                    else:
                        record[field] = value
                else:
                    record[field] = None
            
            processed_data.append(record)
        
        return processed_data
    
    async def _generate_csv_export(
        self,
        data: List[Dict[str, Any]]
    ) -> bytes:
        """Generate CSV export"""
        if not data:
            return b""
        
        output = io.StringIO()
        
        # Get headers from first record
        headers = list(data[0].keys())
        
        writer = csv.DictWriter(output, fieldnames=headers)
        writer.writeheader()
        
        for record in data:
            writer.writerow(record)
        
        return output.getvalue().encode('utf-8')
    
    async def _generate_excel_export(
        self,
        data: List[Dict[str, Any]]
    ) -> bytes:
        """Generate Excel export"""
        if not data:
            return b""
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Personnel Export"
        
        # Define styles
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        header_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )
        
        # Write headers
        headers = list(data[0].keys())
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = header_border
        
        # Write data
        for row_num, record in enumerate(data, 2):
            for col_num, (key, value) in enumerate(record.items(), 1):
                cell = ws.cell(row=row_num, column=col_num, value=value)
                cell.alignment = Alignment(horizontal="left", vertical="center")
                cell.border = Border(
                    left=Side(style="thin"),
                    right=Side(style="thin"),
                    top=Side(style="thin"),
                    bottom=Side(style="thin")
                )
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0])
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except Exception as e:
                    logger.warning(f"Unexpected error: {e}")
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to bytes
        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        return excel_buffer.getvalue()
    
    async def _generate_json_export(
        self,
        data: List[Dict[str, Any]]
    ) -> bytes:
        """Generate JSON export"""
        return json.dumps(data, indent=2, default=str, ensure_ascii=False).encode('utf-8')
    
    async def _generate_pdf_export(
        self,
        data: List[Dict[str, Any]]
    ) -> bytes:
        """Generate PDF export (simplified implementation)"""
        # For now, return a simple text-based PDF
        # In a real implementation, you would use a PDF library like ReportLab
        if not data:
            return b""
        
        output = io.StringIO()
        output.write("Personnel Export Report\n")
        output.write("=" * 50 + "\n\n")
        output.write(f"Generated: {datetime.now(timezone.utc).isoformat()}\n")
        output.write(f"Total Records: {len(data)}\n\n")
        
        # Write headers
        if data:
            headers = list(data[0].keys())
            output.write("\t".join(headers) + "\n")
            output.write("-" * (len(headers) * 10) + "\n")
            
            # Write data
            for record in data:
                row_data = []
                for header in headers:
                    value = str(record.get(header, ""))
                    row_data.append(value[:20])  # Truncate long values
                output.write("\t".join(row_data) + "\n")
        
        return output.getvalue().encode('utf-8')
    
    async def get_export_templates(
        self
    ) -> Dict[str, Any]:
        """Get available export templates"""
        return {
            "templates": self.export_templates,
            "field_groups": self.field_groups,
            "formats": self.export_formats
        }
    
    async def get_export_preview(
        self,
        template: str = "BASIC",
        filters: Optional[Dict[str, Any]] = None,
        db: Session = None
    ) -> Dict[str, Any]:
        """
        Get export preview with sample data
        
        Args:
            template: Export template
            filters: Export filters (optional)
            db: Database session
            
        Returns:
            Export preview with sample records
        """
        if db is None:
            db = next(get_db())
        
        # Get sample data (limit to 5 records)
        sample_data = await self._get_filtered_personnel(filters, db)
        sample_data = sample_data[:5]
        
        # Process sample data
        processed_data = await self._process_export_data(
            sample_data, template, None, db
        )
        
        return {
            "template": template,
            "sample_records": processed_data,
            "total_available": len(sample_data),
            "fields": self.field_groups.get(template, []),
            "filters_applied": filters or {}
        }
    
    async def schedule_export(
        self,
        export_config: Dict[str, Any],
        db: Session = None
    ) -> Dict[str, Any]:
        """
        Schedule export job (for large datasets)
        
        Args:
            export_config: Export configuration
            db: Database session
            
        Returns:
            Scheduled export job information
        """
        # This would integrate with a background job system
        # For now, return a mock job ID
        job_id = f"export_job_{datetime.now(timezone.utc).timestamp()}"
        
        return {
            "success": True,
            "job_id": job_id,
            "status": "scheduled",
            "export_config": export_config,
            "estimated_completion": datetime.now(timezone.utc) + timedelta(minutes=5),
            "scheduled_at": datetime.now(timezone.utc).isoformat()
        }
    
    async def get_export_history(
        self,
        limit: int = 50,
        db: Session = None
    ) -> List[Dict[str, Any]]:
        """
        Get export history
        
        Args:
            limit: Maximum number of records to return
            db: Database session
            
        Returns:
            Export history records
        """
        # This would query an export history table
        # For now, return mock data
        return [
            {
                "job_id": f"export_job_{i}",
                "file_name": f"personnel_export_{i}.csv",
                "format": "CSV",
                "template": "BASIC",
                "record_count": 100 + i,
                "status": "completed",
                "created_at": (datetime.now(timezone.utc) - timedelta(hours=i)).isoformat(),
                "completed_at": (datetime.now(timezone.utc) - timedelta(hours=i-1)).isoformat(),
                "created_by": "admin"
            }
            for i in range(1, min(limit, 10))
        ]


# Create singleton instance
personnel_export_service = PersonnelExportService()
