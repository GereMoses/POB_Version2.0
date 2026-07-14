"""
Transport Manifest & Reconciliation API
Helideck / gangway check-in reconciliation for accurate POB tracking.
"""

import asyncio
import io
import logging
from datetime import datetime, date, timedelta
from typing import Any, Dict, List, Optional

from fastapi import APIRouter, Depends, HTTPException, Query, status
from fastapi.responses import Response
from pydantic import BaseModel
from sqlalchemy import func, text
from sqlalchemy.orm import Session, joinedload

from ..core.database import get_db
from ..core.dependencies import get_current_user
from ..models.biotime_models import MusteringEvent, MusteringLog
from ..models.emergency import (
    ManifestEntry,
    Transport,
    TransportSchedule,
)
from ..models.personnel import Personnel

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/transport", tags=["Transport Manifest"])


# ─── Pydantic schemas ──────────────────────────────────────────────────────────

class FlightCreate(BaseModel):
    transport_identifier: str          # e.g. "ZS-HEL1" — created if not found
    transport_type: int = 3            # TransportType enum: 3 = HELICOPTER, 4 = VESSEL
    transport_operator: Optional[str] = None
    transport_capacity: int = 12
    transport_max_payload_kg: Optional[float] = None   # weight-and-balance limit (kg)

    schedule_type: str = "CHARTER"     # REGULAR, CHARTER, STANDBY
    departure_location: str
    arrival_location: str
    departure_time: datetime
    arrival_time: Optional[datetime] = None
    notes: Optional[str] = None

    # Recurring generation (e.g. daily/weekly crew-change rotations)
    repeat_frequency: Optional[str] = None   # DAILY | WEEKLY
    repeat_count: int = 1                     # number of occurrences incl. the first (cap 60)


class ManifestEntryCreate(BaseModel):
    passenger_name: str
    direction: str = "INBOUND"         # INBOUND / OUTBOUND
    emp_code: Optional[str] = None
    company: Optional[str] = None
    id_number: Optional[str] = None
    personnel_id: Optional[int] = None
    body_weight: Optional[float] = None      # kg
    baggage_weight: Optional[float] = None    # kg
    remarks: Optional[str] = None


class CheckInRequest(BaseModel):
    emp_code: str                      # from a badge / QR scan at the helideck / gangway


class ManifestEntryUpdate(BaseModel):
    status: str                        # CONFIRMED / NO_SHOW / OFFLOADED / MANIFESTED
    remarks: Optional[str] = None
    override_capacity: bool = False    # confirm even if the flight is at seat capacity
    override_compliance: bool = False  # board even if not fit to travel (expired medical/cert)
    override_weight: bool = False      # board even if it exceeds the weight-and-balance limit


class FlightStatusUpdate(BaseModel):
    status: str                        # SCHEDULED / CONFIRMED / CANCELLED / COMPLETED / IN_TRANSIT


class CrewMember(BaseModel):
    name: str
    role: str = "PILOT"                # PILOT / CO_PILOT / CAPTAIN / DRIVER / CREW
    license_no: Optional[str] = None


class CrewUpdate(BaseModel):
    crew: List[CrewMember]


class CargoItem(BaseModel):
    description: str
    weight_kg: float = 0
    type: str = "GENERAL"              # GENERAL / EQUIPMENT / FOOD / DANGEROUS_GOODS
    dangerous_goods: bool = False
    un_number: Optional[str] = None    # UN dangerous-goods number
    remarks: Optional[str] = None


class CargoUpdate(BaseModel):
    items: List[CargoItem]


# ─── Helpers ──────────────────────────────────────────────────────────────────

def _schedule_to_dict(s: TransportSchedule) -> Dict[str, Any]:
    entries = s.manifest_entries or []
    confirmed  = sum(1 for e in entries if e.status == "CONFIRMED")
    no_show    = sum(1 for e in entries if e.status == "NO_SHOW")
    offloaded  = sum(1 for e in entries if e.status == "OFFLOADED")
    manifested = sum(1 for e in entries if e.status == "MANIFESTED")
    inbound    = sum(1 for e in entries if e.direction == "INBOUND")
    outbound   = sum(1 for e in entries if e.direction == "OUTBOUND")

    t = s.transport
    capacity = (t.capacity if t and t.capacity else 0)
    seated = confirmed + manifested  # people who hold/took a seat
    body_kg    = sum((e.body_weight or 0) for e in entries)
    baggage_kg = sum((e.baggage_weight or 0) for e in entries)

    cargo = s.cargo_manifest or []
    cargo_kg = sum((c.get("weight_kg") or 0) for c in cargo if isinstance(c, dict))
    has_dg = any(c.get("dangerous_goods") for c in cargo if isinstance(c, dict))
    crew = s.crew or []

    # Departure delay (minutes) = actual − scheduled, once it has actually departed
    delay_min = None
    if s.actual_departure_time and s.departure_time:
        delay_min = round((s.actual_departure_time - s.departure_time).total_seconds() / 60)

    return {
        "id": s.id,
        "schedule_type": s.schedule_type,
        "departure_location": s.departure_location,
        "arrival_location": s.arrival_location,
        "departure_time": s.departure_time.isoformat() if s.departure_time else None,
        "arrival_time": s.arrival_time.isoformat() if s.arrival_time else None,
        "actual_departure_time": s.actual_departure_time.isoformat() if s.actual_departure_time else None,
        "actual_arrival_time": s.actual_arrival_time.isoformat() if s.actual_arrival_time else None,
        "departure_delay_min": delay_min,
        "status": s.status,
        "transport": {
            "id": t.id,
            "identifier": t.identifier,
            "type": t.type,
            "operator": t.operator,
            "capacity": t.capacity,
        } if t else None,
        "pax_total": len(entries),
        "pax_confirmed": confirmed,
        "pax_no_show": no_show,
        "pax_offloaded": offloaded,
        "pax_manifested": manifested,
        "pax_inbound": inbound,
        "pax_outbound": outbound,
        # Seats & weight-and-balance
        "capacity": capacity,
        "seats_taken": seated,
        "seats_available": max(capacity - seated, 0) if capacity else None,
        "overbooked": bool(capacity and seated > capacity),
        "body_weight_kg": round(body_kg, 1),
        "baggage_weight_kg": round(baggage_kg, 1),
        "cargo_weight_kg": round(cargo_kg, 1),
        "total_weight_kg": round(body_kg + baggage_kg + cargo_kg, 1),
        "max_payload_kg": (t.max_payload_kg if t else None),
        "over_payload": bool(t and t.max_payload_kg and (body_kg + baggage_kg + cargo_kg) > t.max_payload_kg),
        # Crew & cargo
        "crew": crew,
        "crew_count": len(crew),
        "cargo": cargo,
        "cargo_count": len(cargo),
        "has_dangerous_goods": has_dg,
        "is_reconciled": manifested == 0 and len(entries) > 0,
        "created_at": s.created_at.isoformat() if s.created_at else None,
    }


def _entry_to_dict(e: ManifestEntry) -> Dict[str, Any]:
    return {
        "id": e.id,
        "schedule_id": e.schedule_id,
        "passenger_name": e.passenger_name,
        "emp_code": e.emp_code,
        "company": e.company,
        "id_number": e.id_number,
        "direction": e.direction,
        "status": e.status,
        "personnel_id": e.personnel_id,
        "body_weight": e.body_weight,
        "baggage_weight": e.baggage_weight,
        "confirmed_at": e.confirmed_at.isoformat() if e.confirmed_at else None,
        "confirmed_by_id": e.confirmed_by_id,
        "remarks": e.remarks,
        "created_at": e.created_at.isoformat() if e.created_at else None,
    }


def _parse_date(v):
    """Best-effort date parse from mixed JSONB / DB values. None if unparseable."""
    if not v:
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    s = str(v)[:19]
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%Y-%m-%dT%H:%M:%S"):
        try:
            return datetime.strptime(s[:len(fmt) + 2], fmt).date()
        except Exception:
            continue
    try:
        return datetime.fromisoformat(s).date()
    except Exception:
        return None


def travel_compliance(p: Optional[Personnel]) -> Dict[str, Any]:
    """Assess fitness to travel: medical + safety training / certifications.

    status: ok | warning | blocked. `blocked` = an expired critical record (should
    stop boarding); `warning` = missing or soon-to-expire records (allow, but flag).
    Never raises — an unparseable record is simply skipped.
    """
    if p is None:
        return {"status": "warning", "issues": ["Guest / not a registered person — no compliance records"]}

    issues: List[str] = []
    blocked = warning = False
    today = date.today()
    soon = today + timedelta(days=30)

    med = _parse_date(getattr(p, "medical_fitness_date", None))
    if med is None:
        issues.append("No medical certificate on file"); warning = True
    elif med < today:
        issues.append(f"Medical expired {med.isoformat()}"); blocked = True
    elif med <= soon:
        issues.append(f"Medical expires {med.isoformat()}"); warning = True

    def _scan(records) -> bool:
        nonlocal blocked, warning
        items = records if isinstance(records, list) else (records.get("items", []) if isinstance(records, dict) else [])
        found = False
        for it in items:
            if not isinstance(it, dict):
                continue
            exp = _parse_date(it.get("expiry_date") or it.get("valid_until")
                              or it.get("expires") or it.get("expiry") or it.get("valid_to"))
            if exp is None:
                continue
            found = True
            name = it.get("name") or it.get("type") or it.get("title") or "Certification"
            if exp < today:
                issues.append(f"{name} expired {exp.isoformat()}"); blocked = True
            elif exp <= soon:
                issues.append(f"{name} expires {exp.isoformat()}"); warning = True
        return found

    has_training = _scan(getattr(p, "training_records", None))
    has_certs = _scan(getattr(p, "certifications", None))
    if not has_training and not has_certs:
        issues.append("No safety training / certifications on file"); warning = True

    status = "blocked" if blocked else ("warning" if warning else "ok")
    return {"status": status, "issues": issues}


# ─── Flight endpoints ──────────────────────────────────────────────────────────

@router.get("/flights")
async def list_flights(
    date_from: Optional[date] = None,
    date_to: Optional[date] = None,
    flight_status: Optional[str] = Query(None, alias="status"),
    direction: Optional[str] = None,
    db: Session = Depends(get_db),
    _=Depends(get_current_user),
) -> Dict[str, Any]:
    """List flights with optional date/status filters. Defaults to today ± 7 days."""
    if date_from is None:
        date_from = date.today() - timedelta(days=7)
    if date_to is None:
        date_to = date.today() + timedelta(days=7)

    q = (
        db.query(TransportSchedule)
        .options(
            joinedload(TransportSchedule.transport),
            joinedload(TransportSchedule.manifest_entries),
        )
        .filter(
            TransportSchedule.departure_time >= datetime.combine(date_from, datetime.min.time()),
            TransportSchedule.departure_time <= datetime.combine(date_to, datetime.max.time()),
        )
    )
    if flight_status:
        q = q.filter(TransportSchedule.status == flight_status.upper())

    schedules = q.order_by(TransportSchedule.departure_time).all()

    # Optional direction filter — filter schedules that have entries in that direction
    if direction:
        direction_up = direction.upper()
        schedules = [
            s for s in schedules
            if any(e.direction == direction_up for e in (s.manifest_entries or []))
            or not s.manifest_entries  # include empty flights too
        ]

    return {
        "flights": [_schedule_to_dict(s) for s in schedules],
        "total": len(schedules),
        "date_from": str(date_from),
        "date_to": str(date_to),
    }


@router.post("/flights", status_code=status.HTTP_201_CREATED)
async def create_flight(
    payload: FlightCreate,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
) -> Dict[str, Any]:
    """Create a new flight/voyage schedule, auto-creating the transport if not found."""
    transport = (
        db.query(Transport)
        .filter(Transport.identifier == payload.transport_identifier)
        .first()
    )
    if not transport:
        transport = Transport(
            identifier=payload.transport_identifier,
            type=payload.transport_type,
            operator=payload.transport_operator,
            capacity=payload.transport_capacity,
            max_payload_kg=payload.transport_max_payload_kg,
            is_available=True,
        )
        db.add(transport)
        db.flush()

    # Build the departure/arrival times for each occurrence (recurring rotations).
    step = None
    freq = (payload.repeat_frequency or "").upper()
    if freq == "DAILY":
        step = timedelta(days=1)
    elif freq == "WEEKLY":
        step = timedelta(weeks=1)
    occurrences = max(1, min(payload.repeat_count or 1, 60)) if step else 1

    first_schedule = None
    for i in range(occurrences):
        offset = (step * i) if step else timedelta(0)
        schedule = TransportSchedule(
            transport_id=transport.id,
            schedule_type=payload.schedule_type,
            departure_location=payload.departure_location,
            arrival_location=payload.arrival_location,
            departure_time=payload.departure_time + offset,
            arrival_time=(payload.arrival_time + offset) if payload.arrival_time else None,
            frequency=freq or None,
            status="SCHEDULED",
            special_requirements=payload.notes,
        )
        db.add(schedule)
        if first_schedule is None:
            db.flush()
            first_schedule = schedule
    db.commit()

    schedule = (
        db.query(TransportSchedule)
        .options(
            joinedload(TransportSchedule.transport),
            joinedload(TransportSchedule.manifest_entries),
        )
        .filter(TransportSchedule.id == first_schedule.id)
        .one()
    )
    out = _schedule_to_dict(schedule)
    out["occurrences_created"] = occurrences
    return out


@router.get("/flights/{flight_id}")
async def get_flight(
    flight_id: int,
    db: Session = Depends(get_db),
    _=Depends(get_current_user),
) -> Dict[str, Any]:
    schedule = (
        db.query(TransportSchedule)
        .options(
            joinedload(TransportSchedule.transport),
            joinedload(TransportSchedule.manifest_entries),
        )
        .filter(TransportSchedule.id == flight_id)
        .first()
    )
    if not schedule:
        raise HTTPException(status_code=404, detail="Flight not found")
    return _schedule_to_dict(schedule)


@router.patch("/flights/{flight_id}/status")
async def update_flight_status(
    flight_id: int,
    payload: FlightStatusUpdate,
    db: Session = Depends(get_db),
    _=Depends(get_current_user),
) -> Dict[str, Any]:
    schedule = db.query(TransportSchedule).filter(TransportSchedule.id == flight_id).first()
    if not schedule:
        raise HTTPException(status_code=404, detail="Flight not found")

    new_status = payload.status.upper()
    now = datetime.utcnow()
    # Stamp actuals on the real-world transitions so delays can be computed.
    if new_status == "IN_TRANSIT" and not schedule.actual_departure_time:
        schedule.actual_departure_time = now
    if new_status == "COMPLETED" and not schedule.actual_arrival_time:
        schedule.actual_arrival_time = now
        if not schedule.actual_departure_time:  # completed without an explicit depart
            schedule.actual_departure_time = schedule.departure_time

    schedule.status = new_status
    db.commit()
    delay = None
    if schedule.actual_departure_time and schedule.departure_time:
        delay = round((schedule.actual_departure_time - schedule.departure_time).total_seconds() / 60)
    return {
        "id": flight_id,
        "status": schedule.status,
        "actual_departure_time": schedule.actual_departure_time.isoformat() if schedule.actual_departure_time else None,
        "actual_arrival_time": schedule.actual_arrival_time.isoformat() if schedule.actual_arrival_time else None,
        "departure_delay_min": delay,
    }


# ─── Manifest endpoints ────────────────────────────────────────────────────────

@router.get("/flights/{flight_id}/manifest")
async def get_manifest(
    flight_id: int,
    db: Session = Depends(get_db),
    _=Depends(get_current_user),
) -> Dict[str, Any]:
    schedule = db.query(TransportSchedule).filter(TransportSchedule.id == flight_id).first()
    if not schedule:
        raise HTTPException(status_code=404, detail="Flight not found")

    entries = (
        db.query(ManifestEntry)
        .filter(ManifestEntry.schedule_id == flight_id)
        .order_by(ManifestEntry.direction, ManifestEntry.passenger_name)
        .all()
    )

    # Resolve compliance for all registered passengers in one query
    pids = [e.personnel_id for e in entries if e.personnel_id]
    people = {p.id: p for p in db.query(Personnel).filter(Personnel.id.in_(pids)).all()} if pids else {}

    out_entries = []
    for e in entries:
        d = _entry_to_dict(e)
        d["compliance"] = travel_compliance(people.get(e.personnel_id) if e.personnel_id else None)
        out_entries.append(d)

    return {
        "flight_id": flight_id,
        "entries": out_entries,
        "total": len(entries),
        "summary": _schedule_to_dict(schedule),
    }


@router.post("/flights/{flight_id}/manifest", status_code=status.HTTP_201_CREATED)
async def add_manifest_entry(
    flight_id: int,
    payload: ManifestEntryCreate,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
) -> Dict[str, Any]:
    """Add a passenger to the manifest. If emp_code given, auto-fills name from personnel."""
    schedule = db.query(TransportSchedule).filter(TransportSchedule.id == flight_id).first()
    if not schedule:
        raise HTTPException(status_code=404, detail="Flight not found")
    if schedule.status in ("COMPLETED", "CANCELLED"):
        raise HTTPException(status_code=400, detail=f"Cannot add passengers to a {schedule.status} flight")

    personnel_id = payload.personnel_id
    passenger_name = payload.passenger_name.strip()

    # Auto-lookup by emp_code / personnel_id (also used for the compliance check)
    person = None
    if personnel_id:
        person = db.query(Personnel).filter(Personnel.id == personnel_id).first()
    elif payload.emp_code:
        person = db.query(Personnel).filter(Personnel.emp_code == payload.emp_code).first()
        if person:
            personnel_id = person.id
    if person and not passenger_name:
        passenger_name = f"{person.first_name} {person.last_name}".strip()

    entry = ManifestEntry(
        schedule_id=flight_id,
        personnel_id=personnel_id,
        passenger_name=passenger_name,
        emp_code=payload.emp_code,
        company=payload.company,
        id_number=payload.id_number,
        body_weight=payload.body_weight,
        baggage_weight=payload.baggage_weight,
        direction=payload.direction.upper(),
        status="MANIFESTED",
        remarks=payload.remarks,
    )
    db.add(entry)
    db.commit()
    db.refresh(entry)
    out = _entry_to_dict(entry)
    out["compliance"] = travel_compliance(person)
    return out


@router.patch("/flights/{flight_id}/manifest/{entry_id}")
async def update_manifest_entry(
    flight_id: int,
    entry_id: int,
    payload: ManifestEntryUpdate,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
) -> Dict[str, Any]:
    """Confirm arrival, mark no-show, or offload a passenger."""
    entry = (
        db.query(ManifestEntry)
        .filter(ManifestEntry.id == entry_id, ManifestEntry.schedule_id == flight_id)
        .first()
    )
    if not entry:
        raise HTTPException(status_code=404, detail="Manifest entry not found")

    new_status = payload.status.upper()
    allowed = {"MANIFESTED", "CONFIRMED", "NO_SHOW", "OFFLOADED"}
    if new_status not in allowed:
        raise HTTPException(status_code=400, detail=f"status must be one of {allowed}")

    # Capacity enforcement — don't confirm more passengers than seats (override to bypass).
    if new_status == "CONFIRMED" and entry.status != "CONFIRMED" and not payload.override_capacity:
        cap_sched = db.query(TransportSchedule).filter(TransportSchedule.id == flight_id).first()
        cap = cap_sched.transport.capacity if cap_sched and cap_sched.transport else None
        if cap:
            already = (
                db.query(ManifestEntry)
                .filter(
                    ManifestEntry.schedule_id == flight_id,
                    ManifestEntry.status == "CONFIRMED",
                    ManifestEntry.id != entry.id,
                )
                .count()
            )
            if already >= cap:
                raise HTTPException(
                    status_code=409,
                    detail=f"Flight is at capacity ({already}/{cap}). Offload a passenger or override.",
                )

        # Fitness-to-travel — hard-block boarding someone with an EXPIRED medical /
        # safety certificate (status 'blocked'). Warnings never block; expiry does.
        if entry.personnel_id and not payload.override_compliance:
            person = db.query(Personnel).filter(Personnel.id == entry.personnel_id).first()
            comp = travel_compliance(person)
            if comp["status"] == "blocked":
                raise HTTPException(
                    status_code=409,
                    detail=f"Not fit to travel — {'; '.join(comp['issues'])}. Override to board.",
                )

        # Weight & balance — enforce the aircraft/vessel payload limit when one is set.
        if not payload.override_weight:
            sc = db.query(TransportSchedule).options(joinedload(TransportSchedule.transport)).filter(
                TransportSchedule.id == flight_id).first()
            max_kg = getattr(sc.transport, "max_payload_kg", None) if sc and sc.transport else None
            if max_kg:
                confirmed_wt = (
                    db.query(func.coalesce(func.sum(
                        func.coalesce(ManifestEntry.body_weight, 0) + func.coalesce(ManifestEntry.baggage_weight, 0)
                    ), 0))
                    .filter(ManifestEntry.schedule_id == flight_id,
                            ManifestEntry.status == "CONFIRMED",
                            ManifestEntry.id != entry.id)
                    .scalar() or 0
                )
                cargo_wt = sum((c.get("weight_kg") or 0) for c in (sc.cargo_manifest or []) if isinstance(c, dict))
                this_wt = (entry.body_weight or 0) + (entry.baggage_weight or 0)
                projected = float(confirmed_wt) + float(cargo_wt) + float(this_wt)
                if projected > float(max_kg):
                    raise HTTPException(
                        status_code=409,
                        detail=f"Over payload limit — {projected:.0f} kg would exceed {max_kg:.0f} kg. "
                               f"Offload weight or override.",
                    )

    entry.status = new_status
    if payload.remarks is not None:
        entry.remarks = payload.remarks
    if new_status == "CONFIRMED":
        entry.confirmed_at = datetime.utcnow()
        entry.confirmed_by_id = getattr(current_user, "id", None)
        sched = db.query(TransportSchedule).filter(TransportSchedule.id == entry.schedule_id).first()

        # INBOUND arrival: the person is now on board at the destination — add to POB.
        # (Previously only OUTBOUND was handled, so confirmed arrivals never updated POB.)
        if entry.direction == "INBOUND" and entry.personnel_id:
            db.query(Personnel).filter(Personnel.id == entry.personnel_id).update(
                {
                    "is_onboard": True,
                    "is_pob": True,
                    "pob_location": (sched.arrival_location if sched else None),
                    "pob_since": datetime.utcnow(),
                },
                synchronize_session=False,
            )

        # OUTBOUND departure clears zone/POB state and marks the person safe/departed
        # in any active mustering event so they don't appear as MISSING after leaving.
        if entry.direction == "OUTBOUND" and entry.personnel_id:
            db.query(Personnel).filter(Personnel.id == entry.personnel_id).update(
                {"current_zone_id": None, "is_onboard": False, "is_pob": False, "pob_location": None},
                synchronize_session=False,
            )
            if entry.emp_code:
                active_event = (
                    db.query(MusteringEvent).filter(MusteringEvent.status == 0).first()
                )
                if active_event:
                    active_log = (
                        db.query(MusteringLog)
                        .filter(
                            MusteringLog.event_id == active_event.id,
                            MusteringLog.emp_code == entry.emp_code,
                        )
                        .first()
                    )
                    if active_log and active_log.status != 1:
                        from ..services.mustering_service import MusteringService
                        MusteringService(db).mark_person_status(
                            active_event.id,
                            entry.emp_code,
                            status=1,
                            marked_by=getattr(current_user, "id", 0),
                        )
    elif new_status in ("NO_SHOW", "OFFLOADED"):
        entry.confirmed_at = None
        entry.confirmed_by_id = None

    db.commit()
    return _entry_to_dict(entry)


@router.post("/flights/{flight_id}/checkin")
async def checkin_by_badge(
    flight_id: int,
    payload: CheckInRequest,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
) -> Dict[str, Any]:
    """Check a passenger in by scanning their badge / QR (emp_code) at the
    helideck / gangway. Confirms their manifest entry — reusing the same POB and
    mustering side-effects as a manual confirm."""
    emp = (payload.emp_code or "").strip()
    if not emp:
        raise HTTPException(status_code=400, detail="emp_code is required")
    entry = (
        db.query(ManifestEntry)
        .filter(ManifestEntry.schedule_id == flight_id, ManifestEntry.emp_code == emp)
        .order_by(ManifestEntry.id)
        .first()
    )
    if not entry:
        raise HTTPException(status_code=404, detail=f"{emp} is not on this flight's manifest")
    if entry.status == "CONFIRMED":
        return {**_entry_to_dict(entry), "already_checked_in": True}
    # Reuse the confirm path (weight/POB/mustering side-effects all apply).
    result = await update_manifest_entry(
        flight_id, entry.id, ManifestEntryUpdate(status="CONFIRMED"), db, current_user
    )
    result["checked_in"] = True
    result["passenger_name"] = entry.passenger_name
    return result


# ─── Crew & cargo (per-journey) ─────────────────────────────────────────────────

@router.get("/flights/{flight_id}/crew")
async def get_crew(flight_id: int, db: Session = Depends(get_db), _=Depends(get_current_user)):
    s = db.query(TransportSchedule).filter(TransportSchedule.id == flight_id).first()
    if not s:
        raise HTTPException(status_code=404, detail="Flight not found")
    return {"crew": s.crew or []}


@router.put("/flights/{flight_id}/crew")
async def set_crew(flight_id: int, payload: CrewUpdate, db: Session = Depends(get_db),
                   _=Depends(get_current_user)):
    s = db.query(TransportSchedule).filter(TransportSchedule.id == flight_id).first()
    if not s:
        raise HTTPException(status_code=404, detail="Flight not found")
    s.crew = [m.dict() for m in payload.crew]
    db.commit()
    return {"crew": s.crew}


@router.get("/flights/{flight_id}/cargo")
async def get_cargo(flight_id: int, db: Session = Depends(get_db), _=Depends(get_current_user)):
    s = db.query(TransportSchedule).filter(TransportSchedule.id == flight_id).first()
    if not s:
        raise HTTPException(status_code=404, detail="Flight not found")
    items = s.cargo_manifest or []
    return {
        "items": items,
        "total_weight_kg": round(sum((c.get("weight_kg") or 0) for c in items if isinstance(c, dict)), 1),
        "has_dangerous_goods": any(c.get("dangerous_goods") for c in items if isinstance(c, dict)),
    }


@router.put("/flights/{flight_id}/cargo")
async def set_cargo(flight_id: int, payload: CargoUpdate, db: Session = Depends(get_db),
                    _=Depends(get_current_user)):
    s = db.query(TransportSchedule).filter(TransportSchedule.id == flight_id).first()
    if not s:
        raise HTTPException(status_code=404, detail="Flight not found")
    s.cargo_manifest = [i.dict() for i in payload.items]
    db.commit()
    return {"items": s.cargo_manifest}


# ─── Notifications, export & analytics ──────────────────────────────────────────

class NotifyBody(BaseModel):
    subject: Optional[str] = None
    message: Optional[str] = None
    email: bool = True
    sms: bool = False


@router.post("/flights/{flight_id}/notify")
async def notify_passengers(flight_id: int, payload: NotifyBody, db: Session = Depends(get_db),
                            _=Depends(get_current_user)) -> Dict[str, Any]:
    """Notify the manifested passengers of this flight by email / SMS (booking,
    schedule change, delay). Reuses the shared env-configured senders."""
    s = (
        db.query(TransportSchedule)
        .options(joinedload(TransportSchedule.transport), joinedload(TransportSchedule.manifest_entries))
        .filter(TransportSchedule.id == flight_id).first()
    )
    if not s:
        raise HTTPException(status_code=404, detail="Flight not found")

    ident = s.transport.identifier if s.transport else f"#{flight_id}"
    route = f"{s.departure_location} → {s.arrival_location}"
    when = s.departure_time.strftime("%d %b %Y %H:%M") if s.departure_time else "TBC"
    subject = payload.subject or f"Transport {ident}: {route}"
    body = payload.message or (
        f"You are manifested on {ident} ({route}), departing {when}. "
        f"Please be at the departure point in good time with your ID."
    )

    entries = s.manifest_entries or []
    pids = [e.personnel_id for e in entries if e.personnel_id]
    people = {p.id: p for p in db.query(Personnel).filter(Personnel.id.in_(pids)).all()} if pids else {}
    result: Dict[str, Any] = {"email": None, "sms": None}

    if payload.email:
        from ..services.notify_email import send_email, smtp_configured
        emails = sorted({people[e.personnel_id].email for e in entries
                         if e.personnel_id and people.get(e.personnel_id) and people[e.personnel_id].email})
        if not smtp_configured():
            result["email"] = {"sent": 0, "error": "Email not configured (Settings → Email)"}
        else:
            result["email"] = await asyncio.to_thread(send_email, list(emails), subject, f"<p>{body}</p>", body)

    if payload.sms:
        from ..services.notify_sms import send_sms, sms_configured
        phones = sorted({(people[e.personnel_id].phone or people[e.personnel_id].emergency_contact_phone)
                         for e in entries if e.personnel_id and people.get(e.personnel_id)
                         and (people[e.personnel_id].phone or people[e.personnel_id].emergency_contact_phone)})
        if not sms_configured():
            result["sms"] = {"sent": 0, "error": "SMS gateway not configured"}
        else:
            result["sms"] = await asyncio.to_thread(send_sms, list(phones), body)

    return {"success": True, "result": result}


@router.get("/flights/{flight_id}/manifest.xlsx")
async def export_manifest_xlsx(flight_id: int, db: Session = Depends(get_db),
                               _=Depends(get_current_user)):
    """Download a printable Excel manifest for the pilot / captain."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    s = (
        db.query(TransportSchedule)
        .options(joinedload(TransportSchedule.transport), joinedload(TransportSchedule.manifest_entries))
        .filter(TransportSchedule.id == flight_id).first()
    )
    if not s:
        raise HTTPException(status_code=404, detail="Flight not found")

    summary = _schedule_to_dict(s)
    ident = s.transport.identifier if s.transport else f"#{flight_id}"
    wb = Workbook()
    ws = wb.active
    ws.title = "Manifest"
    thin = Side(style="thin", color="FFB0B0B0")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    hdr_fill = PatternFill("solid", fgColor="FF1E3A8A")

    ws.merge_cells("A1:G1")
    ws["A1"] = f"TRANSPORT MANIFEST — {ident}"
    ws["A1"].font = Font(bold=True, size=15, color="FFFFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor="FF111827")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 26

    meta = [
        ("Route", f"{s.departure_location} → {s.arrival_location}"),
        ("Departure", s.departure_time.strftime("%d %b %Y  %H:%M") if s.departure_time else "—"),
        ("Operator", (s.transport.operator if s.transport else "") or "—"),
        ("Crew", ", ".join(f"{c.get('name')} ({c.get('role')})" for c in (s.crew or [])) or "—"),
        ("Weight", f"{summary['total_weight_kg']} kg  (pax {summary['body_weight_kg']} + bag {summary['baggage_weight_kg']} + cargo {summary['cargo_weight_kg']})"),
        ("Seats", f"{summary['seats_taken']}/{summary['capacity'] or '—'}"),
    ]
    r = 2
    for label, val in meta:
        ws.cell(row=r, column=1, value=label).font = Font(bold=True)
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=7)
        ws.cell(row=r, column=2, value=val)
        r += 1

    r += 1
    headers = ["#", "Passenger", "Emp/ID", "Company", "Direction", "Weight (kg)", "Status"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=r, column=c, value=h)
        cell.font = Font(bold=True, color="FFFFFFFF")
        cell.fill = hdr_fill
        cell.border = border
        cell.alignment = Alignment(horizontal="center")
    for i, e in enumerate(sorted(s.manifest_entries or [], key=lambda x: (x.direction, x.passenger_name)), 1):
        r += 1
        wt = (e.body_weight or 0) + (e.baggage_weight or 0)
        row = [i, e.passenger_name, e.emp_code or e.id_number or "", e.company or "",
               e.direction, wt or "", e.status]
        for c, val in enumerate(row, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.border = border

    ws.column_dimensions["B"].width = 26
    ws.column_dimensions["D"].width = 20
    for col in ("A", "C", "E", "F", "G"):
        ws.column_dimensions[col].width = 14

    buf = io.BytesIO()
    wb.save(buf)
    fname = f"manifest_{ident}_{s.departure_time.strftime('%Y%m%d') if s.departure_time else flight_id}.xlsx"
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


@router.get("/analytics")
async def transport_analytics(db: Session = Depends(get_db), _=Depends(get_current_user)) -> Dict[str, Any]:
    """Dashboard metrics: today's activity, on-time %, and a 7-day POB movement trend."""
    today = date.today()
    day_start = datetime.combine(today, datetime.min.time())
    day_end = datetime.combine(today, datetime.max.time())

    todays = db.query(TransportSchedule).filter(
        TransportSchedule.departure_time >= day_start,
        TransportSchedule.departure_time <= day_end,
    ).all()
    today_ids = [f.id for f in todays]
    today_entries = (
        db.query(ManifestEntry).filter(ManifestEntry.schedule_id.in_(today_ids)).all()
        if today_ids else []
    )
    pax_moved = sum(1 for e in today_entries if e.status == "CONFIRMED")

    # On-time %: of flights that actually departed, how many within 15 min of schedule
    departed = db.query(TransportSchedule).filter(
        TransportSchedule.actual_departure_time.isnot(None)
    ).order_by(TransportSchedule.actual_departure_time.desc()).limit(100).all()
    on_time = 0
    for f in departed:
        if f.departure_time and f.actual_departure_time:
            delay = (f.actual_departure_time - f.departure_time).total_seconds() / 60
            if delay <= 15:
                on_time += 1
    on_time_pct = round(on_time / len(departed) * 100) if departed else None

    # 7-day POB movement trend (confirmed arrivals − departures per day)
    trend = []
    for d in range(6, -1, -1):
        day = today - timedelta(days=d)
        ds = datetime.combine(day, datetime.min.time())
        de = datetime.combine(day, datetime.max.time())
        fids = [r[0] for r in db.query(TransportSchedule.id).filter(
            TransportSchedule.departure_time >= ds, TransportSchedule.departure_time <= de).all()]
        arr = dep = 0
        if fids:
            for e in db.query(ManifestEntry).filter(ManifestEntry.schedule_id.in_(fids),
                                                    ManifestEntry.status == "CONFIRMED").all():
                if e.direction == "INBOUND":
                    arr += 1
                else:
                    dep += 1
        trend.append({"date": day.isoformat(), "arrivals": arr, "departures": dep, "net": arr - dep})

    return {
        "flights_today": len(todays),
        "pax_moved_today": pax_moved,
        "on_time_pct": on_time_pct,
        "pob_trend": trend,
    }


@router.delete("/flights/{flight_id}/manifest/{entry_id}", status_code=status.HTTP_204_NO_CONTENT)
async def delete_manifest_entry(
    flight_id: int,
    entry_id: int,
    db: Session = Depends(get_db),
    _=Depends(get_current_user),
):
    entry = (
        db.query(ManifestEntry)
        .filter(ManifestEntry.id == entry_id, ManifestEntry.schedule_id == flight_id)
        .first()
    )
    if not entry:
        raise HTTPException(status_code=404, detail="Manifest entry not found")
    db.delete(entry)
    db.commit()


# ─── POB impact summary ────────────────────────────────────────────────────────

@router.get("/pob-summary")
async def pob_summary(
    summary_date: Optional[date] = None,
    db: Session = Depends(get_db),
    _=Depends(get_current_user),
) -> Dict[str, Any]:
    """POB delta for a given day (defaults to today): confirmed arrivals minus confirmed departures."""
    if summary_date is None:
        summary_date = date.today()

    day_start = datetime.combine(summary_date, datetime.min.time())
    day_end = datetime.combine(summary_date, datetime.max.time())

    flights_today = (
        db.query(TransportSchedule)
        .filter(
            TransportSchedule.departure_time >= day_start,
            TransportSchedule.departure_time <= day_end,
        )
        .all()
    )

    flight_ids = [f.id for f in flights_today]
    if not flight_ids:
        return {
            "date": str(summary_date),
            "confirmed_arrivals": 0,
            "confirmed_departures": 0,
            "no_shows": 0,
            "pob_delta": 0,
            "flights_today": 0,
            "discrepancies": [],
        }

    entries = (
        db.query(ManifestEntry)
        .filter(ManifestEntry.schedule_id.in_(flight_ids))
        .all()
    )

    confirmed_inbound  = [e for e in entries if e.status == "CONFIRMED" and e.direction == "INBOUND"]
    confirmed_outbound = [e for e in entries if e.status == "CONFIRMED" and e.direction == "OUTBOUND"]
    no_shows = [e for e in entries if e.status == "NO_SHOW"]

    # Discrepancies: passengers still MANIFESTED on completed / in-transit flights
    completed_ids = {f.id for f in flights_today if f.status in ("COMPLETED", "IN_TRANSIT")}
    unreconciled = [
        {
            "flight_id": e.schedule_id,
            "passenger_name": e.passenger_name,
            "emp_code": e.emp_code,
            "direction": e.direction,
        }
        for e in entries
        if e.status == "MANIFESTED" and e.schedule_id in completed_ids
    ]

    return {
        "date": str(summary_date),
        "confirmed_arrivals": len(confirmed_inbound),
        "confirmed_departures": len(confirmed_outbound),
        "no_shows": len(no_shows),
        "pob_delta": len(confirmed_inbound) - len(confirmed_outbound),
        "flights_today": len(flights_today),
        "discrepancies": unreconciled,
    }


# ─── Personnel search helper ───────────────────────────────────────────────────

@router.get("/personnel-search")
async def search_personnel(
    q: str = Query(..., min_length=2),
    db: Session = Depends(get_db),
    _=Depends(get_current_user),
) -> List[Dict[str, Any]]:
    """Quick search for adding registered personnel to a manifest."""
    term = f"%{q}%"
    results = (
        db.query(Personnel)
        .filter(
            (Personnel.first_name + " " + Personnel.last_name).ilike(term)
            | Personnel.emp_code.ilike(term)
        )
        .filter(Personnel.status == "active")
        .limit(20)
        .all()
    )
    return [
        {
            "id": p.id,
            "emp_code": p.emp_code,
            "name": f"{p.first_name} {p.last_name}".strip(),
            "company": getattr(p, "company", None),
            "department": p.department if hasattr(p, "department") and isinstance(p.department, str) else None,
        }
        for p in results
    ]


# ─── Mustering reconciliation ──────────────────────────────────────────────────

@router.get("/flights/{flight_id}/reconcile")
async def reconcile_manifest_with_mustering(
    flight_id: int,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
) -> Dict[str, Any]:
    """
    Cross-check the flight manifest against the current live mustering headcount.

    Returns three lists:
    - verified:        manifest passengers confirmed accounted for in mustering
    - missing_in_muster: on manifest but NOT yet checked in to any muster zone
    - extra_on_platform: in live muster but NOT on this manifest (unexpected personnel)

    A departure should only be authorised when missing_in_muster is empty.
    """
    schedule = db.query(TransportSchedule).filter(TransportSchedule.id == flight_id).first()
    if not schedule:
        raise HTTPException(status_code=404, detail="Flight schedule not found")

    # Manifest emp_codes for this flight
    manifest_emp_codes: set = {
        e.emp_code for e in (schedule.manifest_entries or [])
        if e.emp_code and e.status not in ("CANCELLED", "NO_SHOW", "OFFLOADED")
    }

    # Active mustering personnel — anyone with an open muster log entry (status=0=missing or 1=safe)
    muster_rows = db.execute(text("""
        SELECT DISTINCT ml.emp_code
        FROM mustering_log ml
        JOIN mustering_event me ON me.id = ml.event_id
        WHERE me.status = 0   -- active event only
    """)).fetchall()
    muster_emp_codes: set = {r[0] for r in muster_rows}

    # If no active muster, fall back to current POB (last seen on platform)
    if not muster_emp_codes:
        pob_rows = db.execute(text("""
            SELECT DISTINCT p.emp_code
            FROM personnel p
            WHERE p.status = 'ACTIVE'
              AND EXISTS (
                  SELECT 1 FROM iclock_transaction t
                  WHERE t.emp_code = p.emp_code
                    AND t.punch_time >= NOW() - INTERVAL '24 hours'
              )
        """)).fetchall()
        muster_emp_codes = {r[0] for r in pob_rows}
        headcount_source = "pob_24h"
    else:
        headcount_source = "active_muster"

    verified             = sorted(manifest_emp_codes & muster_emp_codes)
    missing_in_muster    = sorted(manifest_emp_codes - muster_emp_codes)
    extra_on_platform    = sorted(muster_emp_codes   - manifest_emp_codes)

    # Enrich missing list with names for the safety officer
    def _name(emp_code: str) -> str:
        row = db.execute(text(
            "SELECT first_name || ' ' || COALESCE(last_name,'') FROM personnel WHERE emp_code=:c LIMIT 1"
        ), {"c": emp_code}).fetchone()
        return row[0].strip() if row else emp_code

    return {
        "flight_id":        flight_id,
        "headcount_source": headcount_source,
        "reconciled":       len(missing_in_muster) == 0,
        "summary": {
            "manifest_total":   len(manifest_emp_codes),
            "verified":         len(verified),
            "missing_in_muster": len(missing_in_muster),
            "extra_on_platform": len(extra_on_platform),
        },
        "missing_in_muster": [
            {"emp_code": c, "name": _name(c)} for c in missing_in_muster
        ],
        "extra_on_platform": [
            {"emp_code": c, "name": _name(c)} for c in extra_on_platform
        ],
        "verified_emp_codes": verified,
        "generated_at": datetime.now().isoformat(),
    }
